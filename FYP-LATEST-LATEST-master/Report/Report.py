import openpyxl
import tkinter as tk
from tkinter import messagebox
from Report.get_keyword_score import get_keyword_score
from Report.get_similarity_score import get_similarity_score
from Report.get_grammar_score import get_grammar_score
from Report.get_spelling_score import get_spelling_score
import utils
from Report.Det_Report import Det_Report

class Report(tk.Frame):  # Using Frame instead of Toplevel
    def __init__(self, master):
        super().__init__(master)
        self.master = master  # Store reference to main window
        self.create_report_ui()

    def create_report_ui(self):
        """Clears the current window and displays the report UI"""
        for widget in self.master.winfo_children():
            widget.destroy()  # Remove previous UI elements

        # Call scoring functions
        get_similarity_score(utils.list_user_input, utils.finalList)
        get_keyword_score(utils.list_user_input, utils.all_keywords)
        get_spelling_score(utils.list_user_input)
        get_grammar_score(utils.list_user_input)

        # Score calculation factors
        similarity_factor = 0.4
        keyword_factor = 0.3
        grammar_factor = 0.15
        spell_factor = 0.15

        Qtext = utils.Qtext
        all_final_score = [''] * len(Qtext)
        
        # Compute total scores
        total = 0
        for i in range(len(Qtext)):
            all_final_score[i] = round(
                utils.all_similarity_scores[i] * similarity_factor +
                utils.all_keyword_scores[i] * keyword_factor +
                utils.all_grammar_scores[i] * grammar_factor +
                utils.all_spelling_scores[i] * spell_factor, 
                2
            )
            if utils.all_keyword_scores[i] <= 0.2:
                all_final_score[i] = 0

        total = round(sum(all_final_score) * 10, 2)
        utils.total = total
        utils.all_final_score = all_final_score

        # Save score to Excel file
        op = "users.xlsx"
        workbook = openpyxl.load_workbook(op)
        sheet = workbook.active
        next_row = sheet.max_row + 1

        # Find user row
        r = next_row
        for i in range(2, next_row):
            if sheet.cell(row=i, column=1).value == utils.user:
                r = i
                break
        
        sheet.cell(row=r, column=get_workbook_column()).value = total
        workbook.save(op)

        # UI elements
        tk.Label(self.master, text=f"Your Marks: {total} out of 100", font=("Arial", 12)).pack(pady=15)
        tk.Button(self.master, text="Detailed Report", width=15, command=self.show_detailed_report).pack(pady=5)
        tk.Button(self.master, text="Back", width=15, command=self.go_back).pack(pady=5)

    def show_detailed_report(self):
        Det_Report(self.master)  # Open detailed report in same window

    def go_back(self):
        """Goes back to the question paper window"""
        from QuestionPaper import QuestionPaper  # Import the QuestionPaper class
        QuestionPaper(self.master)  # Reload question paper UI

def get_workbook_column():
    """Returns the correct column based on the selected subject"""
    subject_columns = {
        "Cryptography": 3,
        "Cyber-Security": 4,
        "E-Commerce": 5,
        "NLP": 6,
        "Philosophy": 7
    }
    return subject_columns.get(utils.file_name, 3)  # Default to column 3
